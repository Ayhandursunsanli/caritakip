# departmanlar/views.py

from django.shortcuts import render, get_object_or_404, redirect
from django.contrib import messages
from .models import Departman
from .forms import DepartmanForm
from satin_alma.models import SatinAlma
from urunler.models import Urun
from tedarikciler.models import Tedarikci
from django.db.models import Sum
from django.http import HttpResponse
import xlsxwriter
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill, numbers
from openpyxl.utils import get_column_letter
from io import BytesIO

def departman_listesi(request):
    departmanlar = Departman.objects.all()
    return render(request, 'departmanlar/departman_listesi.html', {'departmanlar': departmanlar})

def departman_detay(request, pk):
    departman = get_object_or_404(Departman, pk=pk)
    satin_almalar = SatinAlma.objects.filter(departman=departman)

    # Filtreler
    urun_id = request.GET.get('urun')
    tedarikci_id = request.GET.get('tedarikci')
    tarih1 = request.GET.get('tarih1')
    tarih2 = request.GET.get('tarih2')

    if urun_id:
        satin_almalar = satin_almalar.filter(urun_id=urun_id)
    if tedarikci_id:
        satin_almalar = satin_almalar.filter(tedarikci_id=tedarikci_id)
    if tarih1:
        satin_almalar = satin_almalar.filter(tarih__gte=tarih1)
    if tarih2:
        satin_almalar = satin_almalar.filter(tarih__lte=tarih2)

    toplam_tutar = sum([s.toplam_tutar for s in satin_almalar])

    urunler = Urun.objects.all()
    tedarikciler = Tedarikci.objects.all()

    return render(request, 'departmanlar/departman_detay.html', {
        'departman': departman,
        'satin_almalar': satin_almalar,
        'toplam_tutar': toplam_tutar,
        'urunler': urunler,
        'tedarikciler': tedarikciler,
        'filtre': {
            'urun': urun_id,
            'tedarikci': tedarikci_id,
            'tarih1': tarih1,
            'tarih2': tarih2,
        }
    })

def departman_ekle(request):
    if request.method == "POST":
        form = DepartmanForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, 'Departman başarıyla eklendi.')
            return redirect('departmanlar:departman_listesi')
    else:
        form = DepartmanForm()
    return render(request, 'departmanlar/departman_form.html', {'form': form})

def departman_duzenle(request, pk):
    departman = get_object_or_404(Departman, pk=pk)
    if request.method == "POST":
        form = DepartmanForm(request.POST, instance=departman)
        if form.is_valid():
            form.save()
            messages.success(request, 'Departman başarıyla güncellendi.')
            return redirect('departmanlar:departman_listesi')
    else:
        form = DepartmanForm(instance=departman)
    return render(request, 'departmanlar/departman_form.html', {'form': form})

def departman_sil(request, pk):
    departman = get_object_or_404(Departman, pk=pk)
    if request.method == "POST":
        departman.delete()
        messages.success(request, 'Departman başarıyla silindi.')
        return redirect('departmanlar:departman_listesi')
    return render(request, 'departmanlar/departman_sil.html', {'departman': departman})

def departman_excel(request, pk):
    departman = get_object_or_404(Departman, pk=pk)
    satin_almalar = SatinAlma.objects.filter(departman=departman)

    # Filtreler
    urun_id = request.GET.get('urun')
    tedarikci_id = request.GET.get('tedarikci')
    tarih1 = request.GET.get('tarih1')
    tarih2 = request.GET.get('tarih2')

    if urun_id:
        satin_almalar = satin_almalar.filter(urun_id=urun_id)
    if tedarikci_id:
        satin_almalar = satin_almalar.filter(tedarikci_id=tedarikci_id)
    if tarih1:
        satin_almalar = satin_almalar.filter(tarih__gte=tarih1)
    if tarih2:
        satin_almalar = satin_almalar.filter(tarih__lte=tarih2)

    # Excel hazırlık
    wb = Workbook()
    ws = wb.active
    ws.title = "Satın Almalar"

    # Stil ayarları
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="4F81BD")
    alignment_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    alignment_right = Alignment(horizontal="right", vertical="center")
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    headers = [
        "Sıra No", "Tarih", "Departman", "Tedarikçi", "Ürün",
        "Miktar", "Birim", "Fiyat", "Para Birimi", "KDV (%)",
        "KDV Tutarı", "Net Tutar", "Toplam Tutar (KDV Dahil)", "Faturalar"
    ]
    ws.append(headers)

    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = alignment_center
        cell.border = thin_border
        ws.column_dimensions[get_column_letter(col_num)].width = 18

    for row_index, satinalma in enumerate(satin_almalar, start=2):
        miktar = satinalma.miktar or 0
        birim_fiyat = satinalma.birim_fiyat or 0
        kdv_oran = satinalma.kdv.oran if satinalma.kdv else 0
        net_tutar = miktar * birim_fiyat
        kdv_tutar = net_tutar * (kdv_oran / 100)
        toplam_tutar = net_tutar + kdv_tutar

        row_data = [
            row_index - 1,
            satinalma.tarih.strftime('%Y-%m-%d') if satinalma.tarih else '',
            satinalma.departman.ad if satinalma.departman else '',
            satinalma.tedarikci.ad if satinalma.tedarikci else '',
            satinalma.urun.ad if satinalma.urun else '',
            miktar,
            satinalma.birim.ad if satinalma.birim else '',
            birim_fiyat,
            satinalma.para_birimi.kod if satinalma.para_birimi else '',
            kdv_oran,
            kdv_tutar,
            net_tutar,
            toplam_tutar,
            ', '.join([f.fatura_no for f in satinalma.faturalari.all()])
        ]

        for col_num, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_index, column=col_num, value=value)
            cell.border = thin_border

            if col_num in [6, 8, 11, 12, 13]:
                cell.number_format = '#,##0.00'
                cell.alignment = alignment_right
            elif col_num == 10:
                cell.number_format = '0"%"'
                cell.alignment = alignment_center
            else:
                cell.alignment = alignment_center

    # Dosyayı yanıtla
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    response = HttpResponse(
        output.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename=departman_{departman.id}_satinalma_listesi.xlsx'
    return response

# Geçici olarak boş bırakıldı. Gerektiğinde fonksiyonlar eklenecek. 