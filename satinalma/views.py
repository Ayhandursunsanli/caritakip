from django.shortcuts import render, get_object_or_404, redirect
from django.contrib import messages
from django.db.models import Sum, Count, Avg, F, ExpressionWrapper, DecimalField, Q
from django.db.models.functions import TruncMonth, Coalesce
from django.utils import timezone
from datetime import timedelta
from decimal import Decimal
from django.http import HttpResponse
import xlsxwriter
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill, numbers
from openpyxl.utils import get_column_letter
from io import BytesIO
from datetime import datetime
from django.contrib.auth.decorators import login_required
from .models import SatinAlma
from .forms import SatinAlmaForm
from faturalar.models import Fatura
from tedarikciler.models import Tedarikci
from urunler.models import Urun
from departmanlar.models import Departman
import json

def export_to_excel(satin_almalar):
    wb = Workbook()
    ws = wb.active
    ws.title = "Satın Almalar"

    # Stil ayarları
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="4F81BD")
    alignment_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    alignment_right = Alignment(horizontal="right", vertical="center")
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    headers = [
        "Sıra No", "Tarih", "Departman", "Tedarikçi", "Ürün",
        "Miktar", "Birim", "Fiyat", "Para Birimi", "KDV (%)",
        "KDV Tutarı", "Net Tutar", "Toplam Tutar (KDV Dahil)", "Faturalar"
    ]
    ws.append(headers)

    # Başlık stilini uygula
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = alignment_center
        cell.border = thin_border
        ws.column_dimensions[get_column_letter(col_num)].width = 18

    for row_index, satinalma in enumerate(satin_almalar, start=2):
        # Veriler
        miktar = satinalma.miktar or 0
        birim_fiyat = satinalma.birim_fiyat or 0
        kdv_oran = satinalma.kdv.oran if satinalma.kdv else 0
        net_tutar = miktar * birim_fiyat
        kdv_tutar = net_tutar * (kdv_oran / 100)
        toplam_tutar = net_tutar + kdv_tutar

        row_data = [
            row_index - 1,
            satinalma.tarih.strftime('%Y-%m-%d') if satinalma.tarih else '',
            satinalma.departman.ad if satinalma.departman else '',
            satinalma.tedarikci.ad if satinalma.tedarikci else '',
            satinalma.urun.ad if satinalma.urun else '',
            miktar,
            satinalma.birim.ad if satinalma.birim else '',
            birim_fiyat,
            satinalma.para_birimi.kod if satinalma.para_birimi else '',
            kdv_oran,
            kdv_tutar,
            net_tutar,
            toplam_tutar,
            ', '.join([f.fatura_no for f in satinalma.faturalari.all()])
        ]

        for col_num, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_index, column=col_num, value=value)
            cell.border = thin_border

            if col_num in [6, 8, 11, 12, 13]:  # Sayısal değerler
                cell.number_format = '#,##0.00'
                cell.alignment = alignment_right
            elif col_num == 10:  # KDV %
                cell.number_format = '0"%"'
                cell.alignment = alignment_center
            else:
                cell.alignment = alignment_center

    # Excel dosyasını belleğe yaz
    excel_file = BytesIO()
    wb.save(excel_file)
    excel_file.seek(0)
    return excel_file

@login_required
def satin_alma_listesi(request):
    # Filtreleme parametrelerini al
    tedarikci_id = request.GET.get('tedarikci')
    urun_id = request.GET.get('urun')
    departman_id = request.GET.get('departman')
    tarih_baslangic = request.GET.get('tarih_baslangic')
    tarih_bitis = request.GET.get('tarih_bitis')
    durum = request.GET.get('durum')
    arama = request.GET.get('arama', '').strip()
    export = request.GET.get('export')

    # Temel sorgu
    satin_almalar = SatinAlma.objects.select_related(
        'tedarikci', 'urun', 'departman', 'birim', 'para_birimi', 'kdv'
    ).prefetch_related('faturalari')

    # Arama filtresi
    if arama:
        satin_almalar = satin_almalar.filter(
            Q(tedarikci__ad__icontains=arama) |
            Q(urun__ad__icontains=arama) |
            Q(departman__ad__icontains=arama) |
            Q(faturalari__fatura_no__icontains=arama) |
            Q(birim__ad__icontains=arama) |
            Q(para_birimi__kod__icontains=arama)
        ).distinct()

    # Filtreleri uygula
    if tedarikci_id:
        satin_almalar = satin_almalar.filter(tedarikci_id=tedarikci_id)
    if urun_id:
        satin_almalar = satin_almalar.filter(urun_id=urun_id)
    if departman_id:
        satin_almalar = satin_almalar.filter(departman_id=departman_id)
    if tarih_baslangic:
        satin_almalar = satin_almalar.filter(tarih__gte=tarih_baslangic)
    if tarih_bitis:
        satin_almalar = satin_almalar.filter(tarih__lte=tarih_bitis)
    if durum:
        if durum == 'faturali':
            satin_almalar = satin_almalar.filter(faturalari__isnull=False).distinct()
        elif durum == 'faturasiz':
            satin_almalar = satin_almalar.filter(faturalari__isnull=True)
        else:
            satin_almalar = satin_almalar.filter(durum=durum)

    # Excel'e aktarma
    if export == 'excel':
        excel_file = export_to_excel(satin_almalar)
        response = HttpResponse(
            excel_file.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename=satin_alma_listesi_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        return response

    # Her satın alma için tutarları hesapla
    satin_alma_listesi = []
    for satinalma in satin_almalar:
        net_tutar = satinalma.birim_fiyat * satinalma.miktar
        kdv_tutari = net_tutar * (satinalma.kdv.oran / 100) if satinalma.kdv else 0
        toplam_tutar = net_tutar + kdv_tutari
        
        satin_alma_listesi.append({
            'model': satinalma,
            'net_tutar': net_tutar,
            'kdv_tutari': kdv_tutari,
            'toplam_tutar': toplam_tutar
        })

    # Filtreleme için gerekli verileri hazırla
    context = {
        'satin_alma_listesi': satin_alma_listesi,
        'tedarikciler': Tedarikci.objects.all(),
        'urunler': Urun.objects.all(),
        'departmanlar': Departman.objects.all(),
        'filtre': {
            'tedarikci': tedarikci_id,
            'urun': urun_id,
            'departman': departman_id,
            'tarih_baslangic': tarih_baslangic,
            'tarih_bitis': tarih_bitis,
            'durum': durum,
            'arama': arama,
        }
    }
    
    return render(request, 'satin_alma/satin_alma_listesi.html', context)

@login_required
def satin_alma_detay(request, pk):
    satin_alma = get_object_or_404(SatinAlma, pk=pk)
    if request.method == "POST" and request.FILES.get("fis"):
        satin_alma.fis = request.FILES["fis"]
        satin_alma.save()
        messages.success(request, 'Fiş başarıyla yüklendi/değiştirildi.')
        return redirect('satinalma:satin_alma_detay', pk=pk)
    return render(request, 'satin_alma_detay.html', {'satin_alma': satin_alma})

@login_required
def satin_alma_ekle(request):
    if request.method == "POST":
        form = SatinAlmaForm(request.POST, request.FILES)
        if form.is_valid():
            form.save()
            messages.success(request, 'Satın alma başarıyla eklendi.')
            return redirect('satinalma:satin_alma_listesi')
    else:
        form = SatinAlmaForm()
    return render(request, 'satin_alma_form.html', {'form': form})

@login_required
def satin_alma_duzenle(request, pk):
    satin_alma = get_object_or_404(SatinAlma, pk=pk)
    if request.method == "POST":
        form = SatinAlmaForm(request.POST, request.FILES, instance=satin_alma)
        if form.is_valid():
            form.save()
            messages.success(request, 'Satın alma başarıyla güncellendi.')
            return redirect('satinalma:satin_alma_detay', pk=pk)
    else:
        form = SatinAlmaForm(instance=satin_alma)
    return render(request, 'satin_alma_form.html', {'form': form})

@login_required
def satin_alma_sil(request, pk):
    satin_alma = get_object_or_404(SatinAlma, pk=pk)
    if request.method == "POST":
        satin_alma.delete()
        messages.success(request, 'Satın alma başarıyla silindi.')
        return redirect('satinalma:satin_alma_listesi')
    return render(request, 'satin_alma_sil.html', {'satin_alma': satin_alma})

@login_required
def dashboard(request):
    # Son 30 günün tarihi
    son_30_gun = timezone.now() - timedelta(days=30)
    
    # Son 6 ayın başlangıç tarihi
    son_6_ay = timezone.now() - timedelta(days=180)
    
    # Toplam tutar hesaplama (KDV dahil)
    toplam_tutar_hesapla = ExpressionWrapper(
        F('birim_fiyat') * F('miktar') * (1 + (Coalesce(F('kdv__oran'), 0) / 100.0)),
        output_field=DecimalField()
    )
    
    # Genel İstatistikler
    genel_istatistikler = {
        'toplam_satinalma': {
            'sayi': SatinAlma.objects.count(),
            'tutar': SatinAlma.objects.annotate(
                toplam=toplam_tutar_hesapla
            ).aggregate(toplam=Sum('toplam'))['toplam'] or 0
        },
        'toplam_fatura': {
            'sayi': Fatura.objects.count(),
            'tutar': Fatura.objects.aggregate(toplam=Sum('tutar'))['toplam'] or 0
        },
        'aktif_tedarikci': Tedarikci.objects.filter(satinalma__isnull=False).distinct().count(),
        'son_30_gun': {
            'satinalma': SatinAlma.objects.filter(tarih__gte=son_30_gun).annotate(
                toplam=toplam_tutar_hesapla
            ).aggregate(
                sayi=Count('id'),
                tutar=Sum('toplam')
            ),
            'fatura': Fatura.objects.filter(fatura_tarihi__gte=son_30_gun).aggregate(
                sayi=Count('id'),
                tutar=Sum('tutar')
            )
        }
    }

    # Aylık Trend (Son 6 ay)
    aylik_trend = list(SatinAlma.objects.filter(
        tarih__gte=son_6_ay
    ).annotate(
        ay=TruncMonth('tarih'),
        toplam=toplam_tutar_hesapla
    ).values('ay').annotate(
        satinalma_tutar=Sum('toplam'),
        fatura_tutar=Sum('faturalari__tutar')
    ).order_by('ay'))

    # JSON uyumlu hale getir
    for item in aylik_trend:
        if item['ay']:
            item['ay'] = item['ay'].isoformat()
        for key in ['satinalma_tutar', 'fatura_tutar']:
            if item.get(key) is not None:
                item[key] = float(item[key])

    # Tedarikçi Dağılımı (Top 10)
    tedarikci_dagilimi = list(SatinAlma.objects.annotate(
        toplam=toplam_tutar_hesapla
    ).values(
        'tedarikci__ad'
    ).annotate(
        toplam_tutar=Sum('toplam')
    ).order_by('-toplam_tutar')[:10])
    for item in tedarikci_dagilimi:
        if item.get('toplam_tutar') is not None:
            item['toplam_tutar'] = float(item['toplam_tutar'])

    # Departman Dağılımı
    departman_dagilimi = list(SatinAlma.objects.annotate(
        toplam=toplam_tutar_hesapla
    ).values(
        'departman__ad'
    ).annotate(
        toplam_tutar=Sum('toplam')
    ).order_by('-toplam_tutar'))
    for item in departman_dagilimi:
        if item.get('toplam_tutar') is not None:
            item['toplam_tutar'] = float(item['toplam_tutar'])

    # En Çok Satın Alınan Ürünler (Top 10)
    populer_urunler = list(SatinAlma.objects.values(
        'urun__ad'
    ).annotate(
        toplam_miktar=Sum('miktar'),
        toplam_tutar=Sum(toplam_tutar_hesapla)
    ).order_by('-toplam_miktar')[:10])
    for item in populer_urunler:
        if item.get('toplam_miktar') is not None:
            item['toplam_miktar'] = float(item['toplam_miktar'])
        if item.get('toplam_tutar') is not None:
            item['toplam_tutar'] = float(item['toplam_tutar'])

    # Son İşlemler
    son_satinalmalar = SatinAlma.objects.select_related(
        'tedarikci', 'urun', 'departman'
    ).annotate(
        hesaplanan_toplam=toplam_tutar_hesapla
    ).order_by('-tarih')[:5]

    son_faturalar = Fatura.objects.select_related(
        'tedarikci'
    ).order_by('-fatura_tarihi')[:5]

    # Bekleyen Ödemeler
    bekleyen_odemeler = Fatura.objects.filter(
        odeme_tarihi__gte=timezone.now()
    ).select_related('tedarikci').order_by('odeme_tarihi')[:5]

    # 1. Faturaya bağlanmamış satın alımı olan tedarikçiler ve toplamları
    tedarikciler_faturasiz = []
    for tedarikci in Tedarikci.objects.filter(satinalma__faturalari=None).distinct():
        satinalmalar = SatinAlma.objects.filter(tedarikci=tedarikci, faturalari=None)
        toplam = sum(s.toplam_tutar for s in satinalmalar)
        if toplam > 0:
            tedarikciler_faturasiz.append({
                'ad': tedarikci.ad,
                'toplam_faturasiz': toplam
            })
    tedarikciler_faturasiz.sort(key=lambda x: x['toplam_faturasiz'], reverse=True)

    # 2. En çok ödeme yapılan tedarikçiler (faturaya bağlanmış)
    tedarikciler_odeme = []
    for tedarikci in Tedarikci.objects.filter(satinalma__faturalari__isnull=False).distinct():
        satinalmalar = SatinAlma.objects.filter(tedarikci=tedarikci, faturalari__isnull=False)
        toplam = sum(s.toplam_tutar for s in satinalmalar)
        if toplam > 0:
            tedarikciler_odeme.append({
                'ad': tedarikci.ad,
                'toplam_odeme': toplam
            })
    tedarikciler_odeme.sort(key=lambda x: x['toplam_odeme'], reverse=True)
    tedarikciler_odeme = tedarikciler_odeme[:10]

    # 3. En çok satın alım yapılan departmanlar
    departmanlar_toplam = []
    for departman in Departman.objects.all():
        satinalmalar = SatinAlma.objects.filter(departman=departman)
        toplam = sum(s.toplam_tutar for s in satinalmalar)
        if toplam > 0:
            departmanlar_toplam.append({
                'ad': departman.ad,
                'toplam_satinalma': toplam
            })
    departmanlar_toplam.sort(key=lambda x: x['toplam_satinalma'], reverse=True)
    departmanlar_toplam = departmanlar_toplam[:10]

    context = {
        'genel_istatistikler': genel_istatistikler,
        'aylik_trend': json.dumps(aylik_trend, ensure_ascii=False, default=str),
        'tedarikci_dagilimi': json.dumps(tedarikci_dagilimi, ensure_ascii=False, default=str),
        'departman_dagilimi': json.dumps(departman_dagilimi, ensure_ascii=False, default=str),
        'populer_urunler': json.dumps(populer_urunler, ensure_ascii=False, default=str),
        'son_satinalmalar': son_satinalmalar,
        'son_faturalar': son_faturalar,
        'bekleyen_odemeler': bekleyen_odemeler,
        'tedarikciler_faturasiz': tedarikciler_faturasiz,
        'tedarikciler_odeme': tedarikciler_odeme,
        'departmanlar_toplam': departmanlar_toplam,
    }
    
    return render(request, 'dashboard.html', context) 